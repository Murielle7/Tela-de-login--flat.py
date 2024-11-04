import subprocess
import pyautogui
import time
import pandas as pd
import os
import pyperclip
from openpyxl import load_workbook

def close_application(application_name):
    os.system(f"taskkill /f /im {application_name}")

def copiar_texto():
    def triple_click(x, y):
        for _ in range(3):
            pyautogui.click(x, y)
            time.sleep(0.1)

    def press_alt_c():
        pyautogui.hotkey('alt', 'c')

    x, y = 114, 1005
    time.sleep(2)
    triple_click(x, y)

    time.sleep(0.5)
    press_alt_c()

    time.sleep(0.5)
    copied_text = pyperclip.paste()

    if not copied_text:
        copied_text = "SIM"

    print("Texto copiado:", copied_text)
    return copied_text

def atualizar_planilha(numero_guia, copied_text):
    caminho_planilha = r'\\10.0.20.65\Financeira\CAJ\Automação - Fianca\Restituição - Planilha automação.xlsx'
    try:
        workbook = load_workbook(caminho_planilha)
        planilha = workbook.active

        linha_encontrada = False
        for row in range(2, planilha.max_row + 1):
            num_guia_cel = planilha[f'A{row}'].value
            print(f"Verificando linha {row}, valor na célula A{row}: {num_guia_cel}")
            if num_guia_cel is not None and str(num_guia_cel).strip() == str(numero_guia).strip():
                print(f"Encontrado numero_guia na linha {row}, atualizando célula H{row}...")
                planilha[f'H{row}'].value = copied_text
                linha_encontrada = True
                break

        if not linha_encontrada:
            print("Número da guia não encontrado na planilha, ou a célula é nula.")

        workbook.save(caminho_planilha)
    except Exception as e:
        print(f"Erro ao atualizar a planilha: {e}")

def preencher_formulario(numero_guia, numero_proad):
    try:
        executavel = r'C:\tjgossh\ttermpro.exe'
        subprocess.Popen(executavel)
        time.sleep(3)

        pyautogui.typewrite("saj001")
        print("Escreveu 'USUÁRIO'.")
        pyautogui.press("tab")
        print("Apertou 'TAB'.")
        pyautogui.typewrite("caj001")
        print("Escreveu 'SENHA'.")
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)
        pyautogui.hotkey('win', 'up')
        pyautogui.press("i")
        print("Apertou 'i'.")
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        pyautogui.press("1")
        print("Apertou '1'.")
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        pyautogui.typewrite("4353094")
        print("Escreveu 'USUÁRIO'.")
        pyautogui.press("tab")
        print("Apertou 'TAB'.")
        pyautogui.typewrite("4353094")
        print("Escreveu 'SENHA'.")
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        pyautogui.typewrite("x")
        print("Escreveu 'x'.")
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        pyautogui.typewrite("x")
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        pyautogui.typewrite("2")
        print("Escreveu '2'.")
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        pyautogui.typewrite("1")
        print("Escreveu '1'.")
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        pyautogui.typewrite(str(numero_guia))
        pyautogui.press("enter")
    except Exception as e:
        print(f"Erro ao preencher o formulário: {e}")
        close_application("ttermpro.exe")
        raise

def salvar_procedimentos(numero_guia, numero_proad, nome_analista, copied_text):
    try:
        pyautogui.typewrite(str(numero_proad))
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        pyautogui.typewrite("Objeto de Ressarcimento")
        print("Escreveu 'Objeto de Ressarcimento'.")
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        pyautogui.typewrite("s")
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        pyautogui.press("F2")
        print("Apertou 'F2'.")
        pyautogui.press("F2")
        print("Apertou 'F2'.")
        pyautogui.typewrite("1")
        print("Apertou '1'.")
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        pyautogui.typewrite("i")
        print("Apertou 'i'.")
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        pyautogui.typewrite(str(numero_guia))
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")

        # Diretório principal e nova pasta por analista
        main_dir = r"\\10.0.20.65\Financeira\CAJ\Automação - Fianca"
        analista_folder_path = os.path.join(main_dir, nome_analista)
                    
        # Cria a pasta do analista, se não existir
        if not os.path.exists(analista_folder_path):
            os.makedirs(analista_folder_path)
            print(f"Criou a pasta '{analista_folder_path}'.")

        # Cria subpasta para o Número do PROAD do analista 
        proad_folder_path = os.path.join(analista_folder_path, str(numero_proad))

        if not os.path.exists(proad_folder_path):
            os.makedirs(proad_folder_path)
            print(f"Criou a pasta '{proad_folder_path}'.")

        # Simular "Ctrl+P" para abrir a janela de impressão
        pyautogui.hotkey('alt', 'p')
        time.sleep(2)  # Espera a janela de impressão abrir
        print("'Deu Alt + P'.")

        # Pressionar 'Enter' para confirmar o salvamento
        pyautogui.press('enter')
        time.sleep(2)  # Tempo para o processo de salvamento
        print("Clicou no botão 'Enter'.")

        # Copiar o caminho completo do arquivo para a área de transferência
        nome_arquivo = f"Desabilitação de Guia"
        caminho_arquivo = os.path.join(proad_folder_path, nome_arquivo)
        pyperclip.copy(caminho_arquivo)
        pyautogui.hotkey('ctrl', 'a')  # Selecionar a barra de localização do nome do arquivo
        pyautogui.hotkey('ctrl', 'v') 
        print(f"Caminho do arquivo '{caminho_arquivo}' definido.")

        # Pressionar 'Enter' para confirmar o local do arquivo
        pyautogui.press('enter')
        time.sleep(2)  # Tempo para o processo de salvamento
        print("Salvou o arquivo PDF.")

        close_application("ttermpro.exe")

    except Exception as e:
        print(f"Erro ao processar a guia: {e}")
        close_application("ttermpro.exe")

def processar_guia(numero_guia, numero_proad, nome_analista):
    try:
        preencher_formulario(numero_guia, numero_proad)
        copied_text = copiar_texto()
        atualizar_planilha(numero_guia, copied_text)
        salvar_procedimentos(numero_guia, numero_proad, nome_analista, copied_text)
    except Exception as e:
        print(f"Erro ao processar a guia: {e}")

def processar_guia_planilha(caminho_planilha):
    df = pd.read_excel(caminho_planilha)

    if 'Status Guia' not in df.columns:
        df['Status Guia'] = ""

    for index, row in df.iterrows():
        numero_guia = row['Número da Guia']
        numero_proad = row['Número do PROAD']
        nome_analista = row['Nome do Analista']
        processar_guia(numero_guia, numero_proad, nome_analista)
        print(f"Processado GUIA: {numero_guia} - PROAD: {numero_proad} - Analista: {nome_analista}")

caminho_planilha = r'\\10.0.20.65\Financeira\CAJ\Automação - Fianca\Restituição - Planilha automação.xlsx'
processar_guia_planilha(caminho_planilha)
