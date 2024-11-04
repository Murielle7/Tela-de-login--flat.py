import openpyxl
from openpyxl.styles import PatternFill

def pintar_linhas_com_celulas_vazias(arquivo_excel, nome_aba):
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(arquivo_excel)
    
    # Selecionar a planilha pelo nome
    ws = wb[nome_aba]
    
    # Definir a cor de preenchimento verde
    azul_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    
    # Encontrar a última linha com dados
    last_row = ws.max_row

    # Lista para armazenar as linhas que devem ser copiadas
    linhas_para_copiar = []

    # Iterar sobre as linhas e verificar as colunas F, G, H, I
    for row in range(1, last_row + 1):
        has_empty_cell = False
        
        # Verificar células nas colunas F (6), G (7), H (8), I (9)
        for col in [7,8]:
            cell = ws.cell(row=row, column=col)
            
            if cell.value is None or cell.value == "":
                has_empty_cell = True
                break
        
        # Se encontrar pelo menos uma célula vazia, pintar a linha de verde
        if has_empty_cell:
            for col in [7,8]:
                cell = ws.cell(row=row, column=col)
                cell.fill = azul_fill
            
            # Adicionar a linha à lista de linhas a serem copiadas
            linhas_para_copiar.append(row)

    # Criar ou selecionar a planilha 'Planilha2'
    if 'azul' in wb.sheetnames:
        ws_destino = wb['azul']
    else:
        ws_destino = wb.create_sheet(title='azul')

    # Copiar e colar as linhas pintadas de verde na 'Planilha2'
    for i, row_index in enumerate(linhas_para_copiar):
        for col in range(1, ws.max_column + 1):  # Copiar todas as colunas
            source_cell = ws.cell(row=row_index, column=col)
            dest_cell = ws_destino.cell(row=i + 1, column=col)  # Começando a colar na linha 1
            dest_cell.value = source_cell.value  # Copiar o valor da célula

    # Salvar o arquivo Excel modificado
    wb.save(arquivo_excel)

# Uso da função
arquivo_excel = 'S:/CAJ/Servidores/Murielle/Demanda DANILO - ALTERAR VALORES (JUNÇÃO).xlsx'
nome_aba = 'Plan1'
pintar_linhas_com_celulas_vazias(arquivo_excel, nome_aba)
