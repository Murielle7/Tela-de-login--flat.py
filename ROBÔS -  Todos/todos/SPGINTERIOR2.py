import subprocess
import pyautogui
import time
import pandas as pd
import os
import pyperclip
from openpyxl import load_workbook

def close_application(application_name):
    os.system(f"taskkill /f /im {application_name}")

def processar_guia(numero_guia, numero_proad, nome_analista):
    try:
        # Caminho do executável
        executavel = r'C:\tjgossh\ttermpro.exe'

        # Abrir o executável
        subprocess.Popen(executavel)

        # Aguardar um tempo para o programa carregar
        time.sleep(3)  # Ajuste conforme necessário

        # Digitar "saj001"
        pyautogui.typewrite("saj001")
        print("Escreveu 'USUÁRIO'.")
        time.sleep(1)

        # Pressionar a tecla "TAB"
        pyautogui.press("tab")
        print("Apertou 'TAB'.")

        # Digitar "caj001"
        pyautogui.typewrite("caj001")
        print("Escreveu 'SENHA'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)
        pyautogui.hotkey('win', 'up')

        # Pressionar a tecla "i"
        pyautogui.press("i")
        print("Apertou 'i'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Pressionar a tecla "1"
        pyautogui.press("1")
        print("Apertou '1'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Digitar "4353094"
        pyautogui.typewrite("4353094")
        print("Escreveu 'USUÁRIO'.")
        time.sleep(1)

        # Pressionar a tecla "TAB"
        pyautogui.press("tab")
        print("Apertou 'TAB'.")

        # Digitar "4353094"
        pyautogui.typewrite("4353094")
        print("Escreveu 'SENHA'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Digitar "x"
        pyautogui.typewrite("x")
        print("Escreveu 'x'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Digitar "x"
        pyautogui.typewrite("x")
        print("Escreveu 'x'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Digitar Gerenciamento do Sistema
        pyautogui.typewrite("2")
        print("Escreveu '2'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Digitar "Insere inform. ressarcimento guia"
        pyautogui.typewrite("1")
        print("Escreveu '1'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Digitar número da guia
        pyautogui.typewrite(str(numero_guia))
        print("Escreveu 'Número da Guia'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        def triple_click(x, y):
                for _ in range(3):
                    pyautogui.click(188,1022)
                    time.sleep(0.1)

        def press_alt_c():
            pyautogui.hotkey('alt', 'c')


        # Coordenadas do campo
        x, y = 188, 1022  # Substitua com as coordenadas reais do campo desejado

        # Espera para garantir que a janela correta esteja ativa
        time.sleep(2)

        # Realiza o clique triplo no campo
        triple_click(x, y)

        # Aguarda um pequeno intervalo
        time.sleep(0.5)

        # Simula a combinação de teclas 'Alt' + 'C'
        press_alt_c()

        # Aguarda para garantir que o texto foi copiado
        time.sleep(0.5)

        # Lê o conteúdo da área de transferência
        copied_text = pyperclip.paste()

        # Verifica se há texto copiado
        if not copied_text:
            copied_text = "SIM"  # Se não houver texto, atribui "SIM"

        # Imprime o conteúdo copiado
        print("Texto copiado:", copied_text)

        # Caminho para a planilha Excel
        caminho_planilha = r'\\10.0.20.65\Financeira\CAJ\Automação - Fianca\Restituição - Planilha automação.xlsx'
        
        # Carrega a planilha
        workbook = load_workbook(caminho_planilha)
        planilha = workbook.active

        # Encontra a linha que corresponde ao valor de 'numero_guia'
        for row in range(2, planilha.max_row + 1):
            num_guia_cel = planilha[f'A{row}'].value
            if num_guia_cel == numero_guia:
                # Atualiza a célula da coluna "Procedimento Feito" na mesma linha
                planilha[f'H{row}'].value = copied_text
                celula = f'H{row}'
                break

        # Salva as alterações na planilha
        workbook.save(caminho_planilha)

        print(f'Texto colado na célula {celula}')
        
        # Digitar número do PROAD
        pyautogui.typewrite(str(numero_proad))
        print("Escreveu 'Número do PROAD'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Digitar "Objeto de Ressarcimento"
        pyautogui.typewrite("Objeto de Ressarcimento")
        print("Escreveu 'Objeto de Ressarcimento'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Digitar "s"
        pyautogui.typewrite("s")
        print("Escreveu 's'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Pressionar a tecla "F2"
        pyautogui.press("F2")
        print("Apertou 'F2'.")
        time.sleep(1)

        # Pressionar a tecla "F2"
        pyautogui.press("F2")
        print("Apertou 'F2'.")
        time.sleep(1)

        # Digitar "1"
        pyautogui.typewrite("1")
        print("Escreveu '1'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Digitar "i"
        pyautogui.typewrite("i")
        print("Escreveu 'i'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Digitar número da guia
        pyautogui.typewrite(str(numero_guia))
        print("Escreveu 'Número da Guia'.")
        time.sleep(1)

        # Pressionar a tecla "ENTER"
        pyautogui.press("enter")
        print("Apertou 'ENTER'.")
        time.sleep(1)

        # Diretório principal e nova pasta por analista
        main_dir = r"\\10.0.20.65\Financeira\CAJ\Automação - Fianca"
        analista_folder_path = os.path.join(main_dir, nome_analista)
                    
        # Cria a pasta do analista, se não existir
        if not os.path.exists(analista_folder_path):
            os.makedirs(analista_folder_path)
            print(f"Criou a pasta '{analista_folder_path}'.")

        # Cria subpasta para o Número do PROAD do analista 
        proad_folder_path = os.path.join(analista_folder_path, str(numero_proad))

        if not os.path.exists(proad_folder_path):
            os.makedirs(proad_folder_path)
            print(f"Criou a pasta '{proad_folder_path}'.")

        # Simular "Ctrl+P" para abrir a janela de impressão
        pyautogui.hotkey('alt', 'p')
        time.sleep(2)  # Espera a janela de impressão abrir
        print("'Deu Alt + P'.")

        # Pressionar 'Enter' para confirmar o salvamento
        pyautogui.press('enter')
        time.sleep(2)  # Tempo para o processo de salvamento
        print("Clicou no botão 'Enter'.")

        # Copiar o caminho completo do arquivo para a área de transferência
        nome_arquivo = f"Desabilitação de Guia"
        caminho_arquivo = os.path.join(proad_folder_path, nome_arquivo)
        pyperclip.copy(caminho_arquivo)
        pyautogui.hotkey('ctrl', 'a')  # Selecionar a barra de localização do nome do arquivo
        pyautogui.hotkey('ctrl', 'v') 
        print(f"Caminho do arquivo '{caminho_arquivo}' definido.")

        # Pressionar 'Enter' para confirmar o local do arquivo
        pyautogui.press('enter')
        time.sleep(2)  # Tempo para o processo de salvamento
        print("Salvou o arquivo PDF.")

        close_application("ttermpro.exe")

    except Exception as e:
        print(f"Erro ao processar a guia: {e}")
        close_application("ttermpro.exe")

def processar_guia_planilha(caminho_planilha):
    df = pd.read_excel(caminho_planilha)

    # Verificar se a coluna 'Status Guia' existe, se não, inicializá-la
    if 'Status Guia' not in df.columns:
        df['Status Guia'] = ""

    # Iterar sobre cada linha da planilha e processar a guia
    for index, row in df.iterrows():
        numero_guia = row['Número da Guia']
        numero_proad = row['Número do PROAD']
        nome_analista = row['Nome do Analista']
        processar_guia(numero_guia, numero_proad, nome_analista)
        print(f"Processado GUIA: {numero_guia} - PROAD: {numero_proad} - Analista: {nome_analista}")


        # Caminho para a planilha Excel
caminho_planilha = r'\\10.0.20.65\Financeira\CAJ\Automação - Fianca\Restituição - Planilha automação.xlsx'
processar_guia_planilha(caminho_planilha)

