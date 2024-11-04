import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Lista completa de cidades em Goiás
cidades_goias = [
    "Abadia de Goiás", "Abadiânia", "Acreúna", "Adelândia", "Água Fria de Goiás", "Água Limpa",
    "Águas Lindas de Goiás", "Alexânia", "Aloândia", "Alto Horizonte","Alto Paraíso de Goiás",
    "Alvorada do Norte", "Amaralina", "Americano do Brasil", "Amorinópolis", "Anápolis", 
    "Anhanguera", "Anicuns", "Aparecida de Goiânia", "Aparecida do Rio Doce", "Aporé", 
    "Araçu", "Aragarças", "Aragoiânia", "Araguapaz", "Arenópolis", "Aruanã", "Aurilândia",
    "Avelinópolis", "Baliza", "Barro Alto", "Bela Vista de Goiás", "Bom Jardim de Goiás",
    "Bom Jesus de Goiás", "Bonfinópolis", "Bonópolis", "Brazabrantes", "Britânia", 
    "Buriti Alegre", "Buriti de Goiás", "Buritinópolis", "Cabeceiras", "Cachoeira Alta", 
    "Cachoeira de Goiás", "Cachoeira Dourada", "Caçu", "Caiapônia", "Caldas Novas", 
    "Caldazinha", "Campestre de Goiás", "Campinaçu", "Campinorte", "Campo Alegre de Goiás", 
    "Campos Belos", "Campos Verdes", "Carmo do Rio Verde", "Castelândia", "Catalão", 
    "Caturaí", "Cavalcante", "Ceres", "Cezarina", "Chapadão do Céu", "Cidade Ocidental", 
    "Cocalzinho de Goiás", "Colinas do Sul", "Córrego do Ouro", "Corumbá de Goiás", 
    "Corumbaíba", "Cristalina", "Cristianópolis", "Crixás", "Cromínia", "Cumari", "Damianópolis", 
    "Damolândia", "Davinópolis", "Diorama", "Divinópolis de Goiás", "Doverlândia", 
    "Edealina", "Edéia", "Estrela do Norte", "Faina", "Fazenda Nova", "Firminópolis", 
    "Flores de Goiás", "Formosa", "Formoso", "Gameleira de Goiás", "Goianápolis", "Goiandira", 
    "Goianésia", "Goiânia", "Goianira", "Goiás", "Goiatuba", "Gouvelândia", "Guapó", 
    "Guaraíta", "Guarani de Goiás", "Guarinos", "Heitoraí", "Hidrolândia", "Hidrolina", 
    "Iaciara", "Inaciolândia", "Indiara", "Inhumas", "Ipameri", "Ipiranga de Goiás", 
    "Iporá", "Israelândia", "Itaberaí", "Itaguari", "Itaguaru", "Itajá", "Itapaci", 
    "Itapirapuã", "Itapuranga", "Itarumã", "Itauçu", "Itumbiara", "Ivolândia", "Jandaia", 
    "Jaraguá", "Jataí", "Jaupaci", "Jesúpolis", "Joviânia", "Jussara", "Lagoa Santa", 
    "Leopoldo de Bulhões", "Luziânia", "Mairipotaba", "Mambaí", "Mara Rosa", "Marzagão", 
    "Matrinchã", "Maurilândia", "Mimoso de Goiás", "Minaçu", "Mineiros", "Moiporá", 
    "Monte Alegre de Goiás", "Montes Claros de Goiás", "Montividiu", "Montividiu do Norte", 
    "Morrinhos", "Morro Agudo de Goiás", "Mossâmedes", "Mozarlândia",
    "Mundo Novo", "Mutunópolis", "Nazário", "Nerópolis", "Niquelândia", "Nova América", 
    "Nova Aurora", "Nova Crixás", "Nova Glória", "Nova Iguaçu de Goiás", "Nova Roma", 
    "Nova Veneza", "Novo Brasil", "Novo Gama", "Novo Planalto", "Orizona", "Ouro Verde de Goiás", 
    "Ouvidor", "Padre Bernardo", "Palestina de Goiás", "Palmeiras de Goiás", "Palmelo", 
    "Palminópolis", "Panamá", "Paranaiguara", "Paraúna", "Perolândia", "Petrolina de Goiás", 
    "Pilar de Goiás", "Piracanjuba", "Piranhas", "Pirenópolis", "Pires do Rio", "Planaltina", 
    "Pontalina", "Porangatu", "Porteirão", "Portelândia", "Posse", "Professor Jamil", 
    "Quirinópolis", "Rialma", "Rianápolis", "Rio Quente", "Rio Verde", "Rubiataba", 
    "Sanclerlândia", "Santa Bárbara de Goiás", "Santa Cruz de Goiás", "Santa Fé de Goiás", 
    "Santa Helena de Goiás", "Santa Isabel", "Santa Rita do Araguaia", "Santa Rita do Novo Destino", 
    "Santa Rosa de Goiás", "Santa Tereza de Goiás", "Santa Terezinha de Goiás", "Santo Antônio da Barra", 
    "Santo Antônio de Goiás", "Santo Antônio do Descoberto", "São Domingos", "São Francisco de Goiás", 
    "São João d'Aliança", "São João da Paraúna", "São Luís de Montes Belos", "São Luiz do Norte", 
    "São Miguel do Araguaia", "São Miguel do Passa Quatro", "São Patrício", "São Simão", 
    "Senador Canedo", "Serranópolis", "Silvânia", "Simolândia", "Sítio d'Abadia", "Taquaral de Goiás", 
    "Teresina de Goiás", "Terezópolis de Goiás", "Três Ranchos", "Trindade", "Trombas", 
    "Turvânia", "Turvelândia", "Uirapuru", "Uruaçu", "Uruana", "Urutaí", "Valparaíso de Goiás", 
    "Varjão", "Vianópolis", "Vicentinópolis", "Vila Boa", "Vila Propício"
]

# Caminho do arquivo
caminho_arquivo = "C:/Users/mferreirasantos/Documents/demanda DANILO -  Atualizar valores.xlsx"

# Carregar a planilha
wb = load_workbook(caminho_arquivo)
ws = wb['Planilha1']

# Percorrer as linhas e verificar se 'Nome Parte' corresponde a alguma cidade de Goiás
for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
    for cell in row:
        if cell.column_letter == 'B':  # Supondo que 'Nome Parte' está na coluna A
            if cell.value and cell.value in cidades_goias:
                # Pintar a linha de vermelho
                for c in row:
                    c.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# Salvar as alterações
wb.save(caminho_arquivo)
