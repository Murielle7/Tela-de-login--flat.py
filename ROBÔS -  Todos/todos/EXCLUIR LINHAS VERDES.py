import openpyxl
from openpyxl.styles import PatternFill

def remover_linhas_verdes(arquivo_excel, nome_aba):
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(arquivo_excel)
    
    # Selecionar a planilha pelo nome
    ws = wb[nome_aba]
    
    # Definir a cor de preenchimento verde (deve ser a mesma usada para colorir as células)
    verde_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    
    # Encontrar a última linha com dados
    last_row = ws.max_row
    
    # Criar uma lista para armazenar as linhas a serem removidas
    linhas_para_remover = []
    
    # Iterar sobre as linhas e verificar se alguma célula nas colunas F, G, H, I está preenchida com verde
    for row in range(1, last_row + 1):
        has_green_cell = False
        
        # Verificar células nas colunas F (6), G (7), H (8), I (9)
        for col in [19]:
            cell = ws.cell(row=row, column=col)
            
            # Verificar se a célula tem o preenchimento verde
            if cell.fill == verde_fill:
                has_green_cell = True
                break
        
        # Adicionar a linha à lista se pelo menos uma célula especificada tiver o preenchimento verde
        if has_green_cell:
            linhas_para_remover.append(row)
    
    # Remover linhas verdes (começando de baixo para cima para evitar reindexação de linhas)
    for row in reversed(linhas_para_remover):
        ws.delete_rows(row, 1)
    
    # Salvar o arquivo Excel modificado
    wb.save(arquivo_excel)

# Uso da função com o caminho e a aba fornecidos
arquivo_excel = 'S:/CAJ/Servidores/Murielle/Dem. PEDRO- Remessa 22 - boletos.xlsx'
nome_aba = 'Planilha1'
remover_linhas_verdes(arquivo_excel, nome_aba)
