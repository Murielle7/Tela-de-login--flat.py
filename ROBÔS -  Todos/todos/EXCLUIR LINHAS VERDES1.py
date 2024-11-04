import openpyxl
from openpyxl.styles import PatternFill

def excluir_linhas_com_celulas_verdes(arquivo_excel, nome_aba):
    try:
        # Carregar o arquivo Excel
        wb = openpyxl.load_workbook(arquivo_excel)
        
        # Selecionar a planilha pelo nome
        ws = wb[nome_aba]
        
        # Encontrar a última linha com dados
        last_row = ws.max_row

        # Cria uma lista para armazenar as linhas a serem excluídas
        linhas_a_excluir = []

        # Iterar sobre as linhas e verificar colunas F, G, H, I
        for row in range(1, last_row + 1):
            linha_tem_verde = False
            
            # Verificar células nas colunas F (6), G (7), H (8), I (9)
            for col in [6, 7, 8, 9]:
                cell = ws.cell(row=row, column=col)
                
                # Comparar a cor de preenchimento da célula
                if cell.fill.start_color.index == '00FF00':
                    linha_tem_verde = True
                    break
                    
            # Se a linha tem pelo menos uma célula verde, marca para exclusão
            if linha_tem_verde:
                linhas_a_excluir.append(row)
        
        # Excluir as linhas de forma invertida para não perder a referência
        for row in reversed(linhas_a_excluir):
            ws.delete_rows(row)

        # Salvar o arquivo Excel modificado
        wb.save(arquivo_excel)
        
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

# Uso da função
arquivo_excel = 'S:/CAJ/Servidores/Murielle/Demanda DANILO - ALTERAR VALORES (JUNÇÃO).xlsx'
nome_aba = 'Plan1'
excluir_linhas_com_celulas_verdes(arquivo_excel, nome_aba)
