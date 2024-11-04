import time
import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import pyautogui
import pyperclip
from openpyxl import load_workbook

# Função para ler o arquivo Excel e retornar todos os valores da coluna especificada
def get_all_values_from_column(file_path, column_name):
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        if column_name in df.columns:
            return df[column_name].dropna().tolist(), df
        else:
            print(f"A coluna '{column_name}' não foi encontrada no arquivo Excel.")
            return [], df
    except Exception as e:
        print(f"Erro ao ler o arquivo Excel: {e}")
        return [], None

# Define tempo de espera para operações de automação
def t(seconds):
    time.sleep(seconds)

# Função para processar uma linha do Excel
def process_line(driver, df, index, value, numero_proad, analista, column_name_status, file_path):
    try:
        driver.get('https://projudi.tjgo.jus.br/PendenciaPublica')
        t(2)  # Espera a página carregar

        # Diretório principal e nova pasta por analista
        main_dir = r"\\10.0.20.65\Financeira\CAJ\Automação - Fianca"
        analista_folder_path = os.path.join(main_dir, analista)

        # Cria a pasta do analista, se não existir
        if not os.path.exists(analista_folder_path):
            os.makedirs(analista_folder_path)
            print(f"Criou a pasta '{analista_folder_path}'.")

        # Cria subpasta para o Número do PROAD do analista
        proad_folder_path = os.path.join(analista_folder_path, str(numero_proad))

        if not os.path.exists(proad_folder_path):
            os.makedirs(proad_folder_path)
            print(f"Criou a pasta '{proad_folder_path}'.")

        # Preenche o campo "Código Localizador" com o valor da coluna
        driver.find_element(By.XPATH, '//*[@id="codPublicacao"]').send_keys(str(value))
        print(f"Dados da coluna 'Validação Alvará' inseridos no campo codPublicacao.")

        # Aguardando um pouco para visualização
        t(5)

        # Clicar no botão "Localizar"
        driver.find_element(By.XPATH, '//*[@id="formLocalizar"]/button').click()
        print("Clicou no botão 'Localizar'.")

        t(5)

        # Trocar para a segunda aba se existir
        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[1])

        # Leitura do HTML e troca de abas
        print("Lendo HTML")

        # Inicializar o estado com "ERRO" como padrão
        status = "ERRO"

        # Tentar ler o texto na segunda aba
        try:
            texto = driver.find_element(By.XPATH, '//*[contains(text(), "não foi possível localizar")]').text
            if "não foi possível localizar" not in texto:
                status = "FEITO"
        except Exception as e:
            # Caso não encontre o texto "não foi possível localizar", assume que o status é "FEITO"
            status = "FEITO"
            print(f"Erro ao ler o texto na segunda aba: {e}")

        # Atualizar o DataFrame com o status
        df.at[index, column_name_status] = str(status)

        # Salve o status atualizado de volta no arquivo Excel, sem afetar a formatação
        try:
            # Carregar o workbook existente
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Atualizar o valor na célula correspondente
            # À medida que a indexação do openpyxl é baseada em 1 (primeira linha é header)
            sheet.cell(row=index + 2, column=column_name_status, value=status)

            # Salvar de volta ao arquivo
            workbook.save(file_path)
            print(f"Arquivo Excel '{file_path}' atualizado com sucesso.")
        except Exception as e:
            print(f"Erro ao salvar o arquivo Excel: {e}")

        # Simular "Ctrl+P" para abrir a janela de impressão
        pyautogui.hotkey('ctrl', 'p')
        t(2)  # Esperar um pouco para a janela de impressão abrir
        print("'Deu Ctrl + P'.")

        # Pressionar 'Enter' para confirmar o salvamento
        pyautogui.press('enter')
        t(2)

        # Copiar o caminho completo do arquivo para a área de transferência
        nome_arquivo = "Validacao Alvara"
        caminho_arquivo = os.path.join(proad_folder_path, nome_arquivo)
        pyperclip.copy(caminho_arquivo)
        pyautogui.hotkey('ctrl', 'a')  # Selecionar a barra de localização do nome do arquivo
        pyautogui.hotkey('ctrl', 'v')
        print(f"Caminho do arquivo '{caminho_arquivo}' definido.")

        # Pressionar 'Enter' para confirmar o local do arquivo
        pyautogui.press('enter')
        t(2)  # Tempo para o processo de salvamento
        print("Salvou o arquivo PDF.")

    except Exception as e:
        print(f"Erro durante a automação: {e}")
        df.at[index, column_name_status] = str(e)  # Escreve o erro no DataFrame

def main():
    # Caminho do arquivo Excel e nome das colunas
    file_path = r"\\10.0.20.65\Financeira\CAJ\Automação - Fianca\Restituição - Planilha automação.xlsx"
    column_name_validacao = "Validação Alvará"
    column_name_proad = "Número do PROAD"
    column_name_analista = "Nome do Analista"
    column_name_status = "Status Alvará"

    # Obtém todos os valores das colunas especificadas do arquivo Excel
    validacoes, df = get_all_values_from_column(file_path, column_name_validacao)
    proads, _ = get_all_values_from_column(file_path, column_name_proad)
    analistas, _ = get_all_values_from_column(file_path, column_name_analista)

    # Adiciona a coluna 'Status Alvará' ao DataFrame se não existir
    # Se ela já existe, convertê-la para string para evitar desformatação não é mais necessário
    if column_name_status not in df.columns:
        df[column_name_status] = ""

    # Converte o mapeamento do nome da coluna para índice da coluna
    def column_name_to_index(name, df):
        return df.columns.get_loc(name) + 1  # 1-based index for Excel

    # Pega o índice da coluna Status Alvará
    column_status_idx = column_name_to_index(column_name_status, df)

    print(f"Linhas a processar: {len(validacoes)}")

    # Itera através de todos os valores para validação a partir da primeira linha
    for index, (value, numero_proad, analista) in enumerate(zip(validacoes[:], proads[:], analistas[:]), start=0):
        driver = None
        try:
            driver = webdriver.Chrome()
            driver.implicitly_wait(10)

            print(f"Processando linha {index} - Value: {value}, PROAD: {numero_proad}, Analista: {analista}")
            process_line(driver, df, index, value, numero_proad, analista, column_status_idx, file_path)
        except Exception as e:
            print(f"Erro ao processar a linha {index}: {e}")
        finally:
            # Fechar o navegador para garantir que ele será fechado após cada iteração
            if driver:
                driver.quit()


if __name__ == "__main__":
    main()
