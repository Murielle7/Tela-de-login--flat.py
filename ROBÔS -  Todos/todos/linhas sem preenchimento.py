import openpyxl
from openpyxl.styles import PatternFill

def pintar_linhas_com_celulas_vazias(arquivo_excel, nome_aba):
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(arquivo_excel)
    
    # Selecionar a planilha pelo nome
    ws = wb[nome_aba]
    
    # Definir a cor de preenchimento verde
    verde_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    
    # Encontrar a última linha com dados
    last_row = ws.max_row
    
    # Iterar sobre as linhas e verificar as colunas S
    for row in range(1, last_row + 1):
        has_empty_cell = False
        
        # Verificar células nas colunas S (18)
        for col in [19]:
            cell = ws.cell(row=row, column=col)
            
            if cell.value is None or cell.value == "":
                has_empty_cell = True
                break
        
        # Se encontrar pelo menos uma célula vazia, pintar a linha de verde
        if has_empty_cell:
            for col in [19]:
                cell = ws.cell(row=row, column=col)
                cell.fill = verde_fill

    # Salvar o arquivo Excel modificado
    wb.save(arquivo_excel)

# Uso da função
arquivo_excel = 'S:/CAJ/Servidores/Murielle/Dem. PEDRO- Remessa 22 - boletos e status.xlsx'
nome_aba = 'Planilha1'
pintar_linhas_com_celulas_vazias(arquivo_excel, nome_aba)
