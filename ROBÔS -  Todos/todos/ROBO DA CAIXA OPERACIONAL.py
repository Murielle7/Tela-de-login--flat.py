import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.keys import Keys
import time
import pyperclip
from pynput.mouse import Button, Controller
from openpyxl import load_workbook

def configurar_driver():
    """
    Configura o driver do Chrome usando webdriver_manager.
    """
    options = Options()
    options.add_argument('--start-maximized')
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
    return driver

def ler_dados_planilha(caminho_arquivo):
    try:
        df = pd.read_excel(caminho_arquivo, sheet_name='06-2024', header=17)  # header=17 para a linha 17 como cabeçalho
        return df
    except Exception as e:
        print(f"Erro ao ler os dados da planilha: {e}")
        return None

def inserir_datas_e_consultar(driver, data_desejada):
    """
    Insere a data desejada nos campos apropriados e realiza a consulta.
    """
    try:
        campo_data_inicial = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="j_id49:filtroView:j_id1458:dtInicialInputDate"]'))
        )
        campo_data_final = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="j_id49:filtroView:j_id1458:dtFinalInputDate"]'))
        )

        campo_data_inicial.send_keys(Keys.CONTROL + "a")
        campo_data_inicial.send_keys(Keys.CONTROL + "v")
        campo_data_inicial.send_keys(Keys.ENTER)
        time.sleep(10)

        campo_data_final.send_keys(Keys.CONTROL + "a")
        campo_data_final.send_keys(Keys.CONTROL + "v")
        campo_data_final.send_keys(Keys.ENTER)
        time.sleep(1)

        botao_consultar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="j_id49:filtroView:j_id1458:j_id1503"]'))
        )
        botao_consultar.click()
    except (TimeoutException, StaleElementReferenceException) as e:
        print(f"Erro ao inserir datas ou realizar a consulta: {e}")

def salvar_dados_planilha(caminho_arquivo_excel, aba, indice, coluna, valor):
    """
    Atualiza o arquivo Excel preservando a formatação existente e adicionando uma coluna se necessário.
    """
    try:
        # Carregar a planilha existente
        wb = load_workbook(caminho_arquivo_excel)
        ws = wb[aba]

        # Encontrar o índice da coluna correta, ou adicionar a coluna se não existir
        col_idx = None
        col_name = None
        for cell in ws[18]:  # Considerando que o cabeçalho está na linha 18 (indice 17)
            if cell.value == coluna:
                col_idx = cell.column
                col_name = cell.column_letter
                break

        if not col_idx:
            col_idx = ws.max_column + 1
            ws.cell(row=18, column=col_idx, value=coluna)

        # Atualizar a célula correta (indice + 19 para definir corretamente a linha, já que o índice do loop começa em 0)
        ws.cell(row=indice + 19, column=col_idx, value=valor)

        # Salvar as alterações
        wb.save(caminho_arquivo_excel)
        print(f"Número '{valor}' inserido com sucesso na linha {indice + 19}, coluna {col_name or col_idx}.")
    except Exception as e:
        print(f"Erro ao salvar os dados na planilha: {e}")

# Caminho para o arquivo Excel
caminho_arquivo = r'\\10.0.20.65\Financeira\CAJ\Transparência\Saldo das Contas Penas Pecuniárias\Contas Caixa Penas Pecuniárias para publicação Portal da Transparência.xlsx'
nome_aba = '06-2024'
linha_nomes_colunas = 17  # Linha 17 como cabeçalho

# Carregar o arquivo Excel
df = pd.read_excel(caminho_arquivo, sheet_name=nome_aba, header=linha_nomes_colunas)
print(df.columns)

for indice, linha in df.iterrows():
    if pd.isna(linha['Agência']):
        print(f"Pular a linha {indice + 1} porque o campo 'Agência' está vazio.")
        continue

    driver = configurar_driver()

    try:
        driver.get("https://depositojudicial.caixa.gov.br/sigsj_internet/login.xhtml")
        wait = WebDriverWait(driver, 20)

        usuario_field = wait.until(EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'usuario') or contains(@name, 'usuario') or contains(@id, 'username') or contains(@name, 'username')]")))
        usuario_field.clear()
        usuario_field.send_keys("FINANCEIRA2024")

        senha_field = wait.until(EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'senha') or contains(@name, 'senha') or contains(@id, 'password') or contains(@name, 'password')]")))
        senha_field.clear()
        senha_field.send_keys("caj2024")
        senha_field.submit()

        numero_agencia = str(int(linha['Agência'])).strip()  # Garantimos que lemos corretamente o valor sem zeros extras
        conta = linha['CONTA']
        dv = linha['DV']
        operacao = "040 - Depósitos Judiciais da Justiça Estadual"

        try:
            campo_agencia = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="j_id49:filtroView:j_id50:agencia"]')))
            campo_agencia.clear()
            campo_agencia.send_keys(str(numero_agencia))
            print('Preencheu ("Agência")')
            campo_conta = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="j_id49:filtroView:j_id50:conta"]')))
            campo_conta.clear()
            campo_conta.send_keys(str(conta))
            print('Preencheu ("CONTA")')
            campo_dv = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="j_id49:filtroView:j_id50:dv"]')))
            campo_dv.clear()
            campo_dv.send_keys(str(dv))
            print('Preencheu ("DV")')
            campo_operacao = driver.find_element(By.XPATH, '//*[@id="j_id49:filtroView:j_id50:operacaoTrue"]')
            campo_operacao.send_keys("040 - Depósitos Judiciais da Justiça Estadual")
            print('Preencheu o campo Operação com: 040 - Depósitos Judiciais da Justiça Estadual')
            botao_consultar = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="j_id49:filtroView:j_id50:btConsultaConta"]')))
            botao_consultar.click()
            print('Clicou no botão Consultar')
            time.sleep(3)
        except Exception as e:
            print(f"Erro ao selecionar a operação ou clicar no botão Consultar: {e}")

        time.sleep(3)

        try:
            imagem_extrato = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//img[@title="Extrato por Período"]')))
            imagem_extrato.click()
            print('Clicou na imagem de Extrato por Período')
        except Exception as e:
            print(f"Erro ao clicar na imagem de 'Extrato por Período': {e}")

        try:
            botao_periodo = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="j_id49:filtroView:j_id1458:icPadraoPesquisa"]/tbody/tr[2]/td/span/span/img'))
            )
            botao_periodo.click()
            print('Clicou no botão Período')
        except Exception as e:
                    print(f"Erro ao clicar no botao 'período': {e}")
        # Inicializa o controlador do mouse
        mouse = Controller()

        # Função para mover o mouse para uma posição e clicar
        def mover_e_clicar(posicao):
            mouse.position = posicao
            time.sleep(0.5)  # Espera meio segundo para garantir que o movimento foi concluído
            mouse.click(Button.left, 1)
            print(f"Clique realizado na posição {posicao}")

        # Lista de posições
        posicoes = [(802, 437), (742, 459)]

        # Percorre todas as posições e realiza os cliques
        for posicao in posicoes:
            mover_e_clicar(posicao)
            time.sleep(0.5)  # Espera meio segundo antes de ir para a próxima posição

        print("Sequência de cliques realizada.")
        data_desejada = "30/06/2024"
        pyperclip.copy(data_desejada)

        inserir_datas_e_consultar(driver, data_desejada)
        try:
            campo_lista = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="j_id49:filtroView:j_id1458:listaView:lista:0:j_id1540"]'))
            )
            texto_campo = campo_lista.text
            print("Conteúdo do campo:")
            print(texto_campo)
        except Exception as e:
            print(f"Erro ao extrair texto do campo: {e}")

        caminho_arquivo_excel = r'\\10.0.20.65\Financeira\CAJ\Transparência\Saldo das Contas Penas Pecuniárias\Contas Caixa Penas Pecuniárias para publicação Portal da Transparência.xlsx'
        aba = "06-2024"
        coluna = "Saldo em 30/06/2024"

        salvar_dados_planilha(caminho_arquivo_excel, aba, indice, coluna, texto_campo)

        print(f'Dados coletados e salvos na planilha na aba "{aba}" na coluna "{coluna}": {texto_campo}')

    except Exception as e:
        print(f"Erro geral no processo para a agência {numero_agencia}: {e}")

    finally:
        driver.quit()

    print(f"Processo para a agência {numero_agencia} concluído.")
